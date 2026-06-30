from datetime import datetime, time
from decimal import Decimal
from io import BytesIO

from django.db.models import Sum, Count, F, Q, Avg
from django.db.models.functions import Coalesce
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from orders.models import Order, OrderItem
from payments.models import Payment


def _parse_date(value):
    if not value:
        return None
    try:
        return timezone.make_aware(datetime.strptime(value, "%Y-%m-%d"))
    except ValueError:
        return None


def get_filtered_orders(request):
    """Aplica filtros comunes sobre pedidos."""
    qs = Order.objects.prefetch_related('items__product', 'payments')

    campaign_slug = request.query_params.get('campaign')
    date_from = _parse_date(request.query_params.get('date_from'))
    date_to = _parse_date(request.query_params.get('date_to'))
    status = request.query_params.get('status')

    if campaign_slug:
        qs = qs.filter(campaign__slug=campaign_slug)
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from.date())
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to.date())
    if status in dict(Order.STATUS_CHOICES):
        qs = qs.filter(status=status)

    return qs.order_by('-created_at')


def _to_decimal(value):
    if value is None:
        return Decimal('0.00')
    if not isinstance(value, Decimal):
        value = Decimal(str(value))
    return value.quantize(Decimal('0.01'))


def _fmt(value):
    """Convierte un valor a string con 2 decimales garantizados."""
    return str(_to_decimal(value))


def build_dashboard_data(qs):
    non_cancelled = qs.exclude(status=Order.STATUS_CANCELLED)
    collected_qs = non_cancelled.filter(status__in=[Order.STATUS_PAID, Order.STATUS_DELIVERED])

    total_sales = _to_decimal(non_cancelled.aggregate(total=Coalesce(Sum('total'), Decimal('0')))['total'])
    total_collected = _to_decimal(collected_qs.aggregate(total=Coalesce(Sum('total'), Decimal('0')))['total'])

    status_counts = {
        status: qs.filter(status=status).count()
        for status, _ in Order.STATUS_CHOICES
    }

    latest_orders = qs[:5].values(
        'code', 'customer_name', 'total', 'status', 'created_at'
    )

    return {
        'total_orders': qs.count(),
        'total_sales': _fmt(total_sales),
        'total_collected': _fmt(total_collected),
        'pending_amount': _fmt(total_sales - total_collected),
        'pending_deliveries': qs.filter(status=Order.STATUS_PAID).count(),
        'orders_by_status': status_counts,
        'latest_orders': list(latest_orders),
    }


def build_campaign_report(qs):
    grouped = (
        qs.values('campaign__id', 'campaign__name', 'campaign__slug')
        .annotate(
            total_orders=Count('id'),
            total_sales=Coalesce(
                Sum('total', filter=~Q(status=Order.STATUS_CANCELLED)), Decimal('0')
            ),
            total_collected=Coalesce(
                Sum('total', filter=Q(status__in=[Order.STATUS_PAID, Order.STATUS_DELIVERED])), Decimal('0')
            ),
        )
        .annotate(pending_amount=F('total_sales') - F('total_collected'))
        .order_by('campaign__name')
    )

    return [
        {
            'id': row['campaign__id'],
            'name': row['campaign__name'],
            'slug': row['campaign__slug'],
            'total_orders': row['total_orders'],
            'total_sales': _fmt(row['total_sales']),
            'total_collected': _fmt(row['total_collected']),
            'pending_amount': _fmt(row['pending_amount']),
        }
        for row in grouped
    ]


def build_product_report(qs):
    items = (
        OrderItem.objects.filter(order__in=qs)
        .exclude(order__status=Order.STATUS_CANCELLED)
        .select_related('product', 'product__campaign')
        .values('product__id', 'product__name', 'product__campaign__name')
        .annotate(
            total_quantity=Sum('quantity'),
            total_revenue=Sum('subtotal'),
            average_price=Avg('unit_price'),
        )
        .order_by('-total_revenue')
    )

    return [
        {
            'id': row['product__id'],
            'name': row['product__name'],
            'campaign': row['product__campaign__name'],
            'total_quantity': row['total_quantity'] or 0,
            'total_revenue': _fmt(row['total_revenue']),
            'average_price': _fmt(row['average_price']),
        }
        for row in items
    ]


def build_classroom_report(qs):
    grouped = (
        qs.values('classroom')
        .annotate(
            total_orders=Count('id'),
            total_amount=Coalesce(
                Sum('total', filter=~Q(status=Order.STATUS_CANCELLED)), Decimal('0')
            ),
        )
        .order_by('-total_amount')
    )

    return [
        {
            'classroom': row['classroom'],
            'total_orders': row['total_orders'],
            'total_amount': _fmt(row['total_amount']),
        }
        for row in grouped
    ]


def export_excel(data):
    """Genera un archivo Excel en memoria a partir de los datos de reporte."""
    wb = Workbook()
    # Hoja Dashboard
    ws = wb.active
    ws.title = "Dashboard"
    dashboard = data['dashboard']
    ws.append(["Métrica", "Valor"])
    ws.append(["Total de pedidos", dashboard['total_orders']])
    ws.append(["Ventas totales", dashboard['total_sales']])
    ws.append(["Cobros verificados", dashboard['total_collected']])
    ws.append(["Monto pendiente", dashboard['pending_amount']])
    ws.append(["Entregas pendientes", dashboard['pending_deliveries']])
    ws.append([])
    ws.append(["Estado", "Cantidad"])
    for st, count in dashboard['orders_by_status'].items():
        ws.append([st, count])

    # Hoja Campañas
    ws = wb.create_sheet("Campañas")
    ws.append(["Campaña", "Pedidos", "Ventas", "Cobrado", "Pendiente"])
    for row in data['campaigns']:
        ws.append([row['name'], row['total_orders'], float(row['total_sales']),
                   float(row['total_collected']), float(row['pending_amount'])])

    # Hoja Productos
    ws = wb.create_sheet("Productos")
    ws.append(["Producto", "Campaña", "Cantidad", "Ingresos", "Precio promedio"])
    for row in data['products']:
        ws.append([row['name'], row['campaign'], row['total_quantity'],
                   float(row['total_revenue']), float(row['average_price'])])

    # Hoja Salas
    ws = wb.create_sheet("Salas")
    ws.append(["Sala", "Pedidos", "Monto total"])
    for row in data['classrooms']:
        ws.append([row['classroom'], row['total_orders'], float(row['total_amount'])])

    # Hoja Pedidos recientes
    ws = wb.create_sheet("Últimos pedidos")
    ws.append(["Código", "Cliente", "Total", "Estado", "Fecha"])
    for row in dashboard['latest_orders']:
        ws.append([row['code'], row['customer_name'], float(row['total']),
                   row['status'], row['created_at'].strftime("%Y-%m-%d %H:%M")])

    # Ajustes visuales básicos
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_pdf(data):
    """Genera un archivo PDF en memoria a partir de los datos de reporte."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Reporte Cooperadora Online", styles['Title']))
    elements.append(Spacer(1, 12))

    dashboard = data['dashboard']
    elements.append(Paragraph("Dashboard", styles['Heading2']))
    summary = [
        ["Total pedidos", str(dashboard['total_orders'])],
        ["Ventas totales", f"$ {dashboard['total_sales']}"],
        ["Cobros verificados", f"$ {dashboard['total_collected']}"],
        ["Monto pendiente", f"$ {dashboard['pending_amount']}"],
        ["Entregas pendientes", str(dashboard['pending_deliveries'])],
    ]
    table = Table(summary, colWidths=[200, 200])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Resumen por campaña", styles['Heading2']))
    rows = [["Campaña", "Pedidos", "Ventas", "Cobrado", "Pendiente"]]
    for row in data['campaigns']:
        rows.append([row['name'], str(row['total_orders']), f"$ {row['total_sales']}",
                     f"$ {row['total_collected']}", f"$ {row['pending_amount']}"])
    table = Table(rows, colWidths=[150, 60, 80, 80, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Resumen por sala", styles['Heading2']))
    rows = [["Sala", "Pedidos", "Monto total"]]
    for row in data['classrooms']:
        rows.append([row['classroom'], str(row['total_orders']), f"$ {row['total_amount']}"])
    table = Table(rows, colWidths=[200, 80, 120])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer
